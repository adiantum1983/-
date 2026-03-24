import streamlit as st
import pandas as pd
import io
from cashflow_direct import load_trial_balance, create_direct_cf_statement

st.set_page_config(page_title="Matrix CF Generator", layout="centered")

st.title("📊 マトリックス法 キャッシュフロージェネレーター")
st.markdown("お手元の **残高試算表 (Excel)** をアップロードするだけで、直接法キャッシュフロー計算書を自動生成します。")

uploaded_file = st.file_uploader("残高試算表のExcelファイルをドラッグ＆ドロップ（または選択）", type=["xlsx", "xls"])

if uploaded_file is not None:
    st.info("ファイルを読み込み、マトリックス法による計算を行っています...")
    try:
        # ファイルパスの代わりに BytesIO のラッパーを渡しても pandas.read_excel はパース可能
        tb_df = load_trial_balance(uploaded_file)
        df_cf = create_direct_cf_statement(tb_df)
        
        st.success("計算完了！以下のプレビューで確認し、Excelファイルとして保存できます。")
        
        st.subheader("生成されたキャッシュフロー計算書（プレビュー）")
        st.dataframe(df_cf, use_container_width=True, hide_index=True)
        
        # ダウンロード用のExcel生成
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_cf.to_excel(writer, index=False, sheet_name="キャッシュフロー計算書", header=["項目", "金額"])
            worksheet = writer.sheets["キャッシュフロー計算書"]
            
            from openpyxl.styles import Font, Border, Side, PatternFill
            thin = Side(border_style="thin", color="000000")
            header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            
            worksheet.column_dimensions['A'].width = 50
            worksheet.column_dimensions['B'].width = 20
            
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=2):
                for cell in row:
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = Font(bold=True)
                    elif cell.col_idx == 2 and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
                    if isinstance(row[0].value, str):
                        val = row[0].value
                        if "I." in val or "II." in val or "III." in val or "小計" in val or "増減額" in val or "期末残高" in val:
                             row[0].font = Font(bold=True)
                             row[1].font = Font(bold=True)
                             
        st.download_button(
            label="⬇️ Excel形式でダウンロード",
            data=buffer.getvalue(),
            file_name="CashFlow_DirectStatement.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        import traceback
        st.error(f"データの処理中にエラーが発生しました: {str(e)}")
        with st.expander("エラー詳細 (開発者向け)"):
            st.text(traceback.format_exc())
            st.write("※A列に「勘定科目コード」(1101, 1122等) が正しく付与された残高試算表かどうか再度ご確認ください。")
